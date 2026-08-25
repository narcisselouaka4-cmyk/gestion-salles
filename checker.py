"""
Module de vérification de disponibilité des salles.
Logique métier complète pour lire les fichiers Excel et Google Sheets.
"""

import re
import os
import base64
import json
import time as time_module
from datetime import datetime, time, date
from openpyxl import load_workbook

# Import pour Google Sheets
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# Mapping des jours
DAY_MAPPING = {
    "lundi": "lundi",
    "mardi": "mardi",
    "mercredi": "mercredi",
    "jeudi": "jeudi",
    "vendredi": "vendredi",
    "vendredi (nuit)": "vendredi",
    "vendredi (soir)": "vendredi",
    "samedi": "samedi",
    "dimanche": "dimanche",
}

PYTHON_WEEKDAY_TO_FR = {
    0: "lundi",
    1: "mardi",
    2: "mercredi",
    3: "jeudi",
    4: "vendredi",
    5: "samedi",
    6: "dimanche"
}

SALLE_KEYWORDS = {
    "salle principale": ["principale", "principal"],
    "salle du fond": ["fond"],
    "salle du milieu": ["milieu"],
}


def normalize_day(cell_value: str) -> str:
    """Normalise une valeur de cellule vers un jour de la semaine."""
    if not cell_value:
        return ""
    val = str(cell_value).lower().strip()
    for key, day in DAY_MAPPING.items():
        if key in val:
            return day
    return val.split()[0] if val else ""


def nth_occurrence_in_month(d: date) -> int:
    """Retourne le rang de ce jour dans le mois (1er lundi = 1, 2ème lundi = 2, etc.)."""
    return (d.day - 1) // 7 + 1


def parse_horaire(horaire_str: str):
    """
    Parse une chaîne horaire et retourne (debut, fin).
    Gère les formats : "15H30 - 18H00", "22H-06H", "09H-12H", etc.
    """
    if not horaire_str:
        raise ValueError("Horaire vide")

    horaire_str = str(horaire_str).upper().replace(" ", "")
    parts = re.split(r"[-–]", horaire_str, maxsplit=1)
    if len(parts) != 2:
        raise ValueError(f"Horaire non parseable : {horaire_str}")

    def parse_time(s: str) -> time:
        s = s.strip().replace("H", ":")
        if s.endswith(":"):
            s += "00"
        if ":" not in s:
            s += ":00"
        h, m = s.split(":")
        return time(int(h), int(m) if m else 0)

    return parse_time(parts[0]), parse_time(parts[1])


def time_in_range(t, debut: time, fin: time) -> bool:
    """Vérifie si l'heure t est dans le créneau [debut, fin], avec gestion overnight."""
    if t is None or not isinstance(t, time):
        return False
    if debut <= fin:
        return debut <= t <= fin
    else:
        # Overnight : le créneau va de debut jusqu'à minuit, puis de minuit jusqu'à fin
        return t >= debut or t <= fin


def salle_matches(cell_value: str, target_salle: str) -> bool:
    """Vérifie si une valeur de cellule correspond à une salle cible."""
    if not cell_value:
        return False
    salle_norm = str(cell_value).lower().strip()
    keywords = SALLE_KEYWORDS.get(target_salle, [target_salle])
    return any(kw in salle_norm for kw in keywords)


def parse_dates(date_cell):
    """
    Parse une ou plusieurs dates depuis une cellule.
    Retourne une liste de dates.
    """
    dates = []

    if date_cell is None:
        return dates

    # Si c'est déjà un datetime
    if isinstance(date_cell, datetime):
        return [date_cell.date()]

    # Si c'est une date
    if isinstance(date_cell, date):
        return [date_cell]

    # Si c'est une chaîne
    date_str = str(date_cell).strip()

    # Cas des multi-dates séparées par "et" ou ","
    separators = ['et', ',', 'et ', ', ']
    parts = [date_str]
    for sep in separators:
        new_parts = []
        for p in parts:
            new_parts.extend([x.strip() for x in p.split(sep) if x.strip()])
        parts = new_parts

    for part in parts:
        # Essayer différents formats
        formats_to_try = [
            '%d/%m/%y',    # 28/02/25
            '%d/%m/%Y',    # 28/02/2025
            '%Y-%m-%d',    # 2025-02-28
        ]

        for fmt in formats_to_try:
            try:
                dt = datetime.strptime(part.strip(), fmt)
                dates.append(dt.date())
                break
            except ValueError:
                continue

    return dates


class SalleChecker:
    """Classe principale pour vérifier la disponibilité des salles."""

    def __init__(self, salles_dir: str, google_sheet_id: str = None):
        self.salles_dir = salles_dir
        self.salles_files = {
            "salle principale": f"{salles_dir}/SALLE PRINCIPALE.xlsx",
            "salle du fond": f"{salles_dir}/SALLE DU FOND.xlsx",
            "salle du milieu": f"{salles_dir}/Salle du Milieu.xlsx",
        }

        # ID du Google Sheet pour les réservations (optionnel)
        self.google_sheet_id = google_sheet_id
        self._google_client = None

        # Cache pour les fichiers de salles (stables)
        self._salles_cache = {}
        # Cache pour le mapping couleur -> occupant (stables)
        self._color_map_cache = {}
        # Cache TTL pour les valeurs Google Sheets (evite les quotas 429)
        self._sheet_cache = {}
        self._sheet_cache_time = {}
        self._sheet_cache_ttl = 15  # secondes

    def _get_worksheet_values_cached(self, worksheet_name: str):
        """
        Recupere les valeurs d'un worksheet Google Sheets avec cache TTL et retry.
        Retourne (values, error).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return None, "Google Sheets non configure"

        cache_key = f"{self.google_sheet_id}:{worksheet_name}"
        now = time_module.time()

        # Utiliser le cache si valide
        if cache_key in self._sheet_cache:
            cached_time = self._sheet_cache_time.get(cache_key, 0)
            if now - cached_time < self._sheet_cache_ttl:
                return self._sheet_cache[cache_key], None

        try:
            client, error = self._get_google_client()
            if error or not client:
                return None, error or "Client Google Sheets non initialise"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet(worksheet_name)
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.get_worksheet(0)

            # Retry en cas de quota exceeded (429)
            max_retries = 3
            last_error = None
            for attempt in range(max_retries):
                try:
                    values = worksheet.get_all_values()
                    self._sheet_cache[cache_key] = values
                    self._sheet_cache_time[cache_key] = now
                    return values, None
                except gspread.exceptions.APIError as api_err:
                    last_error = api_err
                    err_str = str(api_err)
                    if "429" in err_str or "Quota exceeded" in err_str:
                        wait = 2 ** attempt  # 1s, 2s, 4s
                        print(f"[Google Sheets] Quota exceeded, retrying in {wait}s (attempt {attempt + 1}/{max_retries})")
                        time_module.sleep(wait)
                    else:
                        raise

            return None, f"Erreur lecture Google Sheets apres {max_retries} tentatives: {str(last_error)}"

        except Exception as e:
            return None, f"Erreur lecture Google Sheets: {str(e)}"

    def _invalidate_sheet_cache(self, worksheet_name: str):
        """Invalide le cache d'un worksheet apres modification."""
        cache_key = f"{self.google_sheet_id}:{worksheet_name}"
        self._sheet_cache.pop(cache_key, None)
        self._sheet_cache_time.pop(cache_key, None)

    def _normalize_private_key(self, private_key: str) -> str:
        """
        Normalise une clé privée pour Google auth.
        Gère les cas où les retours à la ligne sont échappés (\\n) dans une variable d'env.
        """
        if not private_key:
            return private_key
        # Remplacer les doubles backslash échappés par de vrais retours à la ligne
        if "\\n" in private_key:
            private_key = private_key.replace("\\n", "\n")
        return private_key

    def _get_google_client(self):
        """
        Initialise et retourne le client Google Sheets.
        Retourne (client, error) où error est None si succès, ou un message d'erreur string.
        """
        if not GSPREAD_AVAILABLE:
            return None, "Module gspread non installé"

        if self._google_client is not None:
            return self._google_client, None

        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        errors = []

        # 1. Streamlit secrets (Streamlit Cloud)
        try:
            import streamlit as st
            if "google_credentials" in st.secrets:
                creds_info = dict(st.secrets["google_credentials"])
                creds_info["private_key"] = self._normalize_private_key(creds_info.get("private_key", ""))
                creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
                self._google_client = gspread.authorize(creds)
                print("[Google Auth] Authentifié via Streamlit secrets")
                return self._google_client, None
        except Exception as e:
            errors.append(f"Streamlit secrets: {str(e)}")

        # 2. GOOGLE_CREDENTIALS_JSON (base64) - Render
        import os
        if os.environ.get("GOOGLE_CREDENTIALS_JSON"):
            try:
                json_b64 = os.environ.get("GOOGLE_CREDENTIALS_JSON")
                json_bytes = base64.b64decode(json_b64)
                creds_info = json.loads(json_bytes)
                creds_info["private_key"] = self._normalize_private_key(creds_info.get("private_key", ""))
                creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
                self._google_client = gspread.authorize(creds)
                print("[Google Auth] Authentifié via GOOGLE_CREDENTIALS_JSON")
                return self._google_client, None
            except Exception as e:
                errors.append(f"GOOGLE_CREDENTIALS_JSON: {str(e)}")

        # 3. Variables d'environnement individuelles (Render)
        if os.environ.get("GOOGLE_CREDENTIALS_TYPE"):
            try:
                private_key = self._normalize_private_key(os.environ.get("GOOGLE_CREDENTIALS_PRIVATE_KEY", ""))
                creds_info = {
                    "type": os.environ.get("GOOGLE_CREDENTIALS_TYPE"),
                    "project_id": os.environ.get("GOOGLE_CREDENTIALS_PROJECT_ID"),
                    "private_key_id": os.environ.get("GOOGLE_CREDENTIALS_PRIVATE_KEY_ID"),
                    "private_key": private_key,
                    "client_email": os.environ.get("GOOGLE_CREDENTIALS_CLIENT_EMAIL"),
                    "client_id": os.environ.get("GOOGLE_CREDENTIALS_CLIENT_ID"),
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                }
                creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
                self._google_client = gspread.authorize(creds)
                print("[Google Auth] Authentifié via variables d'environnement")
                return self._google_client, None
            except Exception as e:
                errors.append(f"Variables d'environnement: {str(e)}")

        # 4. Fichiers secrets Render / locaux
        possible_paths = [
            os.path.join(os.path.dirname(__file__), "google-credentials.json"),
            os.path.join(os.path.dirname(__file__), "gcp_credentials.json"),
            "/etc/secrets/google-credentials.json",
            "/etc/secrets/gcp_credentials.json",
            "/mnt/secrets/google-credentials.json",
            "/mnt/secrets/gcp_credentials.json",
            "/secrets/google-credentials.json",
            "/secrets/gcp_credentials.json",
            "google-credentials.json",
            "gcp_credentials.json",
        ]
        for secret_path in possible_paths:
            if os.path.exists(secret_path):
                try:
                    creds = Credentials.from_service_account_file(secret_path, scopes=scopes)
                    self._google_client = gspread.authorize(creds)
                    print(f"[Google Auth] Authentifié via fichier secret: {secret_path}")
                    return self._google_client, None
                except Exception as e:
                    errors.append(f"Fichier secret {secret_path}: {str(e)}")

        # 5. Fichiers credentials.json locaux
        credentials_paths = [
            os.path.join(os.path.dirname(__file__), "credentials.json"),
            os.path.join(os.path.dirname(__file__), "gcp_credentials.json"),
            "/home/visiteur/projet_salles/credentials.json",
            "/home/visiteur/projet_salles/gcp_credentials.json",
            "credentials.json",
            "gcp_credentials.json",
        ]
        for creds_path in credentials_paths:
            if os.path.exists(creds_path):
                try:
                    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
                    self._google_client = gspread.authorize(creds)
                    print(f"[Google Auth] Authentifié via fichier local: {creds_path}")
                    return self._google_client, None
                except Exception as e:
                    errors.append(f"Fichier local {creds_path}: {str(e)}")

        error_msg = "Aucune méthode d'authentification Google n'a fonctionné. " + " | ".join(errors)
        print(f"[Google Auth] {error_msg}")
        return None, error_msg

    def _ensure_planning_headers(self, worksheet):
        """
        Garantit que les en-têtes de la ligne 5 du 'Planning quotidien' incluent
        la colonne K (Téléphone). Étend la grille si nécessaire et copie le style
        des en-têtes existants.
        """
        try:
            # Étendre la grille si pas assez de colonnes (besoin de 11 = A..K)
            if worksheet.col_count < 11:
                worksheet.add_cols(11 - worksheet.col_count)

            values = worksheet.get_all_values()
            if len(values) < 5:
                return
            header_row = values[4]  # ligne 5 (index 4)
            # Si K5 est vide, on ajoute l'en-tête "Téléphone"
            k_val = header_row[10].strip() if len(header_row) > 10 else ""
            if not k_val:
                worksheet.update(values=[['Téléphone']], range_name='K5')
                # Copier le style de J5 vers K5 pour cohérence visuelle
                try:
                    import gspread_formatting as gf
                    fmt = gf.get_user_entered_format(worksheet, 'J5')
                    if fmt:
                        gf.format_cell_ranges(worksheet, [('K5', fmt)])
                except ImportError:
                    pass
                self._invalidate_sheet_cache("Planning quotidien")
        except Exception as e:
            print(f"[ensure_planning_headers] Erreur: {e}")

    def _ensure_users_headers(self, worksheet):
        """
        Garantit que les en-têtes de l'onglet 'Utilisateurs' incluent la colonne F (email).
        Étend la grille si nécessaire.
        """
        try:
            if worksheet.col_count < 6:
                worksheet.add_cols(6 - worksheet.col_count)

            values = worksheet.get_all_values()
            if not values:
                return
            header_row = values[0]
            f_val = header_row[5].strip() if len(header_row) > 5 else ""
            if not f_val:
                worksheet.update(values=[['email']], range_name='F1')
                self._invalidate_sheet_cache("Utilisateurs")
        except Exception as e:
            print(f"[ensure_users_headers] Erreur: {e}")

    def _load_salle_workbook(self, salle_name: str):
        """
        Charge le workbook d'une salle (sans data_only) et met en cache à la fois
        les données et le mapping couleur -> occupant.
        Retourne (data_rows, occupant_map).
        """
        if salle_name in self._salles_cache and salle_name in self._color_map_cache:
            return self._salles_cache[salle_name], self._color_map_cache[salle_name]

        file_path = self.salles_files.get(salle_name)
        if not file_path:
            return [], {}

        try:
            # Charger sans data_only pour accéder aux styles/couleurs ET aux valeurs brutes
            wb = load_workbook(file_path, data_only=False)
            ws = wb.active

            data = []
            occupant_map = {}
            current_block_name = None
            in_gray_block = False

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                data.append(row)

                # Couleur des cellules N (colonne 14) et O (colonne 15)
                cell_n = ws.cell(row=i + 1, column=14)
                cell_o = ws.cell(row=i + 1, column=15)

                color_n = self._cell_bg_color(cell_n)
                color_o = self._cell_bg_color(cell_o)
                is_gray = color_n == "FF999999" or color_o == "FF999999"

                if is_gray:
                    if not in_gray_block:
                        in_gray_block = True
                        current_block_name = None

                    # Chercher un nom dans N ou O
                    for val in (cell_n.value, cell_o.value):
                        if val is not None:
                            name = str(val).strip()
                            if name and current_block_name is None:
                                current_block_name = name
                                break

                    if current_block_name:
                        occupant_map[i] = current_block_name
                else:
                    in_gray_block = False
                    current_block_name = None

            self._salles_cache[salle_name] = data
            self._color_map_cache[salle_name] = occupant_map
            return data, occupant_map

        except Exception as e:
            print(f"Erreur chargement {file_path}: {e}")
            return [], {}

    def _load_salle_data(self, salle_name: str):
        """Charge les données d'un fichier de salle avec mise en cache."""
        data, _ = self._load_salle_workbook(salle_name)
        return data

    def _build_color_occupant_map(self, salle_name: str) -> dict:
        """Retourne le mapping ligne -> occupant basé sur la couleur des cellules N/O."""
        _, occupant_map = self._load_salle_workbook(salle_name)
        return occupant_map

    def _cell_bg_color(self, cell) -> str:
        """Retourne la couleur de fond d'une cellule sous forme de chaîne RGBA."""
        if cell is None or cell.fill is None:
            return "00000000"
        color = cell.fill.start_color
        if color is None:
            return "00000000"
        rgb = color.rgb
        return str(rgb) if rgb else "00000000"

    def check_fixed_schedule(self, salle_name: str, d: date, time_requested: time) -> list:
        """
        Vérifie le planning fixe d'une salle pour une date et heure données.
        Retourne une liste d'occupations.
        """
        results = []
        rows = self._load_salle_data(salle_name)

        if not rows:
            return results

        # Mapping ligne -> occupant basé sur la couleur des cellules N/O
        occupant_map = self._build_color_occupant_map(salle_name)

        # Déterminer le jour de la semaine
        day_name = PYTHON_WEEKDAY_TO_FR[d.weekday()]

        # Déterminer la semaine
        week_number = nth_occurrence_in_month(d)
        week_col_map = {1: 0, 2: 2, 3: 4, 4: 6}
        WEEK_COLS = [0, 2, 4, 6]

        in_5th_week_section = False

        for i, row in enumerate(rows):
            if not row:
                continue

            # Détecter section 5ème semaine
            if row[0] is not None and str(row[0]).strip().upper().startswith("EN PLUS"):
                in_5th_week_section = True
                continue

            # Occupant déterminé par la couleur des cellules N/O
            occupant = occupant_map.get(i)
            if occupant is None:
                continue

            # Déterminer la cellule jour à vérifier
            day_in_cell = None

            # Vérifier si la ligne a une valeur dans au moins une colonne semaine
            has_any_week_col = any(row[col] is not None for col in WEEK_COLS if col < len(row))

            if has_any_week_col:
                # Entrée régulière → ignorer in_5th_week_section
                if week_number <= 4:
                    col_idx = week_col_map.get(week_number)
                    if col_idx is not None and col_idx < len(row):
                        day_in_cell = row[col_idx]
                else:
                    # Semaine 5 → une entrée régulière ne s'applique pas à la 5ème semaine
                    day_in_cell = None
            else:
                # Ligne sans colonne semaine → section "EN PLUS DES MOIS DE 5 DIMANCHES"
                # Cette section ne concerne que les DIMANCHES de la 5ème semaine
                if in_5th_week_section and week_number == 5 and day_name == "dimanche":
                    day_in_cell = day_name  # Force le match
                else:
                    day_in_cell = None

            if day_in_cell is None:
                continue

            # Normaliser et comparer le jour (si c'est une entrée régulière)
            if has_any_week_col:
                cell_day = normalize_day(str(day_in_cell))
                if cell_day != day_name:
                    continue

            # Extraire activité (colonne I = index 8) et horaire (colonne M = index 12)
            activite = str(row[8]).strip() if len(row) > 8 and row[8] else "Non précisée"
            horaire_str = str(row[12]).strip() if len(row) > 12 and row[12] else None

            if horaire_str is None:
                continue

            try:
                debut, fin = parse_horaire(horaire_str)

                # Gestion spéciale pour les créneaux overnight
                if time_in_range(time_requested, debut, fin):
                    results.append({
                        "occupant": occupant,
                        "activite": activite,
                        "horaire": horaire_str,
                        "debut": debut,
                        "fin": fin,
                        "source": "planning fixe"
                    })

            except ValueError:
                # Horaire non parseable → sécurité : considérer occupé
                results.append({
                    "occupant": occupant,
                    "activite": activite,
                    "horaire": horaire_str,
                    "debut": time(0, 0),
                    "fin": time(23, 59),
                    "source": "planning fixe",
                    "warning": "Horaire non parseable"
                })

        return results

    def _handle_horaire_cell(self, horaire_cell):
        """
        Gère une cellule horaire et retourne (debut, fin, label, is_parseable, is_vide).
        - is_parseable = True si on a pu extraire des heures numériques
        - label = texte à afficher dans l'interface
        - is_vide = True si la cellule est vide/None
        """
        if horaire_cell is None:
            return None, None, "Horaire non indiqué", False, True

        horaire_str = str(horaire_cell).strip()

        if not horaire_str:
            return None, None, "Horaire non indiqué", False, True

        # Tenter le parsing numérique
        try:
            debut, fin = parse_horaire(horaire_str)
            return debut, fin, horaire_str, True, False
        except Exception:
            # Pas parseable mais contient du texte → warning
            return None, None, horaire_str, False, False

    def _get_cell_value(self, row, index, default=""):
        """Récupère la valeur d'une cellule. Retourne '' si vide."""
        if len(row) > index and row[index] is not None:
            val = str(row[index]).strip()
            return val if val else default
        return default

    def check_reservations_google(self, salle_name: str, d: date, time_requested: time) -> tuple:
        """
        Vérifie les réservations depuis Google Sheets.
        Retourne (results, error) où error est None si succès.
        """
        results = []

        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return results, None

        all_values, error = self._get_worksheet_values_cached("Planning quotidien")
        if error:
            return results, error

        try:
            for row_idx, row in enumerate(all_values[5:], start=6):  # Commencer à la ligne 6 (index 5)
                if len(row) < 5:
                    continue

                salle_cell = row[1] if len(row) > 1 else None  # Colonne B
                nom_cell = row[2] if len(row) > 2 else None    # Colonne C
                horaire_cell = row[3] if len(row) > 3 else None  # Colonne D
                date_cell = row[4] if len(row) > 4 else None    # Colonne E
                # Nouvelles colonnes (F, G, H, I, J, K)
                accompte = self._get_cell_value(row, 5)       # Colonne F
                reste = self._get_cell_value(row, 6)          # Colonne G
                prix_location = self._get_cell_value(row, 7)  # Colonne H
                caution = self._get_cell_value(row, 8)        # Colonne I
                # Colonne J = "ajouté par [username]"
                added_by_raw = self._get_cell_value(row, 9)
                added_by = ""
                if added_by_raw.startswith("ajouté par "):
                    added_by = added_by_raw.replace("ajouté par ", "")
                telephone = self._get_cell_value(row, 10)      # Colonne K

                if not salle_cell or not date_cell:
                    continue

                if not salle_matches(str(salle_cell).lower().strip(), salle_name):
                    continue

                dates = self._parse_date_from_sheet(date_cell)
                if d not in dates:
                    continue

                nom = str(nom_cell).strip() if nom_cell else "Inconnu"

                infos_manquantes = []
                if not accompte:
                    infos_manquantes.append("Accompte")
                if not reste:
                    infos_manquantes.append("Reste à payer")
                if not prix_location:
                    infos_manquantes.append("Prix de location")
                if not caution:
                    infos_manquantes.append("Chèque caution ménage")

                debut, fin, label, is_parseable, is_vide = self._handle_horaire_cell(horaire_cell)

                reservation_data = {
                    "occupant": nom,
                    "activite": "Réservation ponctuelle",
                    "horaire": label,
                    "accompte": accompte,
                    "reste_a_payer": reste,
                    "prix_location": prix_location,
                    "caution_menage": caution,
                    "salle": str(salle_cell).strip() if salle_cell else "",
                    "added_by": added_by,
                    "telephone": telephone,
                    "source": "réservation"
                }

                if infos_manquantes:
                    reservation_data["infos_manquantes"] = infos_manquantes

                if is_parseable:
                    if time_in_range(time_requested, debut, fin):
                        reservation_data["debut"] = debut
                        reservation_data["fin"] = fin
                        results.append(reservation_data)
                else:
                    reservation_data["debut"] = time(0, 0)
                    reservation_data["fin"] = time(23, 59)
                    if not is_vide:
                        reservation_data["warning"] = "Horaire non indiqué, veuillez le renseigner"
                    results.append(reservation_data)

        except Exception as e:
            return results, f"Erreur lecture Google Sheets: {str(e)}"

        return results, None

    def _parse_date_from_sheet(self, date_cell):
        """Parse une date depuis Google Sheets (qui renvoie des chaînes)."""
        dates = []

        if not date_cell:
            return dates

        # Si c'est déjà un objet date/datetime (rare avec gspread mais possible)
        if isinstance(date_cell, (datetime, date)):
            return [date_cell.date() if isinstance(date_cell, datetime) else date_cell]

        # Sinon traiter comme une chaîne
        date_str = str(date_cell).strip()

        # Essayer de parser différents formats
        formats_to_try = [
            '%d/%m/%y',
            '%d/%m/%Y',
            '%Y-%m-%d',
            '%d-%m-%Y',
        ]

        for fmt in formats_to_try:
            try:
                dt = datetime.strptime(date_str, fmt)
                dates.append(dt.date())
                return dates
            except ValueError:
                continue

        return dates

    def check_reservations(self, salle_name: str, d: date, time_requested: time) -> tuple:
        """
        Vérifie les réservations extérieures pour une salle/date/heure.
        Retourne (occupations, error) où error est None si succès.
        """
        # Le fallback sur fichier Excel a été supprimé ; les réservations
        # ponctuelles sont désormais lues exclusivement depuis Google Sheets.
        if not self.google_sheet_id:
            return [], "Google Sheets non configuré pour les réservations"

        return self.check_reservations_google(salle_name, d, time_requested)

    def get_all_fixed_occupations(self, salle_name: str, d: date) -> list:
        """
        Récupère TOUTES les occupations fixes d'une journée (sans filtrer par heure).
        """
        results = []
        rows = self._load_salle_data(salle_name)

        if not rows:
            return results

        # Mapping ligne -> occupant basé sur la couleur des cellules N/O
        occupant_map = self._build_color_occupant_map(salle_name)

        day_name = PYTHON_WEEKDAY_TO_FR[d.weekday()]
        week_number = nth_occurrence_in_month(d)
        week_col_map = {1: 0, 2: 2, 3: 4, 4: 6}
        WEEK_COLS = [0, 2, 4, 6]

        in_5th_week_section = False

        for i, row in enumerate(rows):
            if not row:
                continue

            if row[0] is not None and str(row[0]).strip().upper().startswith("EN PLUS"):
                in_5th_week_section = True
                continue

            # Occupant déterminé par la couleur des cellules N/O
            occupant = occupant_map.get(i)
            if occupant is None:
                continue

            day_in_cell = None

            # Vérifier si la ligne a une valeur dans au moins une colonne semaine
            has_any_week_col = any(row[col] is not None for col in WEEK_COLS if col < len(row))

            if has_any_week_col:
                # Entrée régulière
                if week_number <= 4:
                    col_idx = week_col_map.get(week_number)
                    if col_idx is not None and col_idx < len(row):
                        day_in_cell = row[col_idx]
                else:
                    # Semaine 5 (hors dimanche) → pas d'occupation fixe
                    # Seul le dimanche de la 5ème semaine peut avoir des occupations (via section "EN PLUS")
                    day_in_cell = None
            else:
                # Ligne sans colonne semaine → section "EN PLUS DES MOIS DE 5 DIMANCHES"
                # Cette section ne concerne que les DIMANCHES de la 5ème semaine
                if in_5th_week_section and week_number == 5 and day_name == "dimanche":
                    day_in_cell = day_name
                else:
                    day_in_cell = None

            if day_in_cell is None:
                continue

            if has_any_week_col:
                cell_day = normalize_day(str(day_in_cell))
                if cell_day != day_name:
                    continue

            activite = str(row[8]).strip() if len(row) > 8 and row[8] else "Non précisée"
            horaire_str = str(row[12]).strip() if len(row) > 12 and row[12] else None

            if horaire_str is None:
                continue

            try:
                debut, fin = parse_horaire(horaire_str)
                results.append({
                    "occupant": occupant,
                    "activite": activite,
                    "horaire": horaire_str,
                    "debut": debut,
                    "fin": fin,
                    "source": "planning fixe"
                })
            except ValueError:
                results.append({
                    "occupant": occupant,
                    "activite": activite,
                    "horaire": horaire_str,
                    "debut": time(0, 0),
                    "fin": time(23, 59),
                    "source": "planning fixe",
                    "warning": "Horaire non parseable"
                })

        return results

    def get_all_reservations_google(self, salle_name: str, d: date) -> tuple:
        """
        Récupère TOUTES les réservations ponctuelles depuis Google Sheets.
        Retourne (results, error) où error est None si succès.
        """
        results = []

        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return results, None

        all_values, error = self._get_worksheet_values_cached("Planning quotidien")
        if error:
            return results, error

        try:
            for row_idx, row in enumerate(all_values[5:], start=6):
                if len(row) < 5:
                    continue

                salle_cell = row[1] if len(row) > 1 else None
                nom_cell = row[2] if len(row) > 2 else None
                horaire_cell = row[3] if len(row) > 3 else None
                date_cell = row[4] if len(row) > 4 else None
                accompte = self._get_cell_value(row, 5)
                reste = self._get_cell_value(row, 6)
                prix_location = self._get_cell_value(row, 7)
                caution = self._get_cell_value(row, 8)
                added_by_raw = self._get_cell_value(row, 9)
                added_by = ""
                if added_by_raw.startswith("ajouté par "):
                    added_by = added_by_raw.replace("ajouté par ", "")
                telephone = self._get_cell_value(row, 10)      # Colonne K

                if not salle_cell or not date_cell:
                    continue

                if not salle_matches(str(salle_cell).lower().strip(), salle_name):
                    continue

                dates = self._parse_date_from_sheet(date_cell)
                if d not in dates:
                    continue

                nom = str(nom_cell).strip() if nom_cell else "Inconnu"

                infos_manquantes = []
                if not accompte:
                    infos_manquantes.append("Accompte")
                if not reste:
                    infos_manquantes.append("Reste à payer")
                if not prix_location:
                    infos_manquantes.append("Prix de location")
                if not caution:
                    infos_manquantes.append("Chèque caution ménage")

                debut, fin, label, is_parseable, is_vide = self._handle_horaire_cell(horaire_cell)

                reservation_data = {
                    "occupant": nom,
                    "activite": "Réservation ponctuelle",
                    "horaire": label,
                    "accompte": accompte,
                    "reste_a_payer": reste,
                    "prix_location": prix_location,
                    "caution_menage": caution,
                    "salle": str(salle_cell).strip() if salle_cell else "",
                    "added_by": added_by,
                    "telephone": telephone,
                    "source": "réservation"
                }

                if infos_manquantes:
                    reservation_data["infos_manquantes"] = infos_manquantes

                if is_parseable:
                    reservation_data["debut"] = debut
                    reservation_data["fin"] = fin
                    results.append(reservation_data)
                else:
                    reservation_data["debut"] = time(0, 0)
                    reservation_data["fin"] = time(23, 59)
                    if not is_vide:
                        reservation_data["warning"] = "Horaire non indiqué, veuillez le renseigner"
                    results.append(reservation_data)

        except Exception as e:
            return results, f"Erreur lecture Google Sheets (get_all): {str(e)}"

        return results, None

    def update_reservation_google(self, salle_name: str, d: date, occupant: str,
                                   new_data: dict) -> tuple:
        """
        Met à jour une réservation dans Google Sheets.
        Retourne (success, error) où error est None si succès.
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Planning quotidien")
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.get_worksheet(0)

            all_values, error = self._get_worksheet_values_cached("Planning quotidien")
            if error:
                return False, error

            # Chercher la ligne correspondante (clé : nom + date, salle optionnelle)
            for row_idx, row in enumerate(all_values[5:], start=6):
                if len(row) < 5:
                    continue

                salle_cell = row[1] if len(row) > 1 else None
                nom_cell = row[2] if len(row) > 2 else None
                date_cell = row[4] if len(row) > 4 else None

                if not nom_cell or not date_cell:
                    continue

                # Correspondance salle : soit vide, soit match (permet de trouver même si salle effacée)
                if salle_cell:
                    if not salle_matches(str(salle_cell).lower().strip(), salle_name):
                        continue

                dates = self._parse_date_from_sheet(date_cell)
                if d not in dates:
                    continue

                nom = str(nom_cell).strip() if nom_cell else ""
                if nom != occupant:
                    continue

                # Mettre à jour les colonnes — vider d'abord les cellules vides
                clear_ranges = []
                updates = []

                def _add_update(col_letter, key):
                    if key in new_data:
                        val = new_data[key]
                        if val:
                            updates.append({'range': f'{col_letter}{row_idx}', 'values': [[val]]})
                        else:
                            clear_ranges.append(f'{col_letter}{row_idx}')

                _add_update('B', 'salle')
                _add_update('C', 'occupant')
                _add_update('D', 'horaire')
                _add_update('E', 'date')
                _add_update('F', 'accompte')
                _add_update('G', 'reste_a_payer')
                _add_update('H', 'prix_location')
                _add_update('I', 'caution_menage')
                _add_update('K', 'telephone')
                # NOTE: colonne J = "ajouté par [username]" — ne pas modifier

                if clear_ranges:
                    worksheet.batch_clear(clear_ranges)

                if updates:
                    worksheet.batch_update(updates)

                self._invalidate_sheet_cache("Planning quotidien")
                return True, None

            return False, "Réservation non trouvée"

        except Exception as e:
            return False, f"Erreur mise à jour Google Sheets: {str(e)}"

    def delete_reservation_google(self, salle_name: str, d: date, occupant: str) -> tuple:
        """
        Supprime une ligne de réservation dans Google Sheets.
        Retourne (success, error).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Planning quotidien")
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.get_worksheet(0)

            all_values, error = self._get_worksheet_values_cached("Planning quotidien")
            if error:
                return False, error

            for row_idx, row in enumerate(all_values[5:], start=6):
                if len(row) < 5:
                    continue

                salle_cell = row[1] if len(row) > 1 else None
                nom_cell = row[2] if len(row) > 2 else None
                date_cell = row[4] if len(row) > 4 else None

                if not nom_cell or not date_cell:
                    continue

                salle_str = str(salle_cell).strip() if salle_cell else ""
                nom_str = str(nom_cell).strip() if nom_cell else ""

                if salle_cell:
                    if not salle_matches(str(salle_cell).lower().strip(), salle_name):
                        continue

                dates = self._parse_date_from_sheet(date_cell)
                if d not in dates:
                    continue

                if nom_str != occupant:
                    continue

                worksheet.delete_rows(row_idx)
                self._invalidate_sheet_cache("Planning quotidien")
                return True, None

            return False, "Réservation non trouvée"

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, f"Erreur suppression Google Sheets: {str(e)}"

    def add_reservation_google(self, data: dict) -> tuple:
        """
        Ajoute une nouvelle réservation dans Google Sheets à la première ligne vide.
        Retourne (success, error_or_info).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Planning quotidien")
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.get_worksheet(0)

            # S'assurer que la colonne K (Téléphone) existe et a son en-tête
            self._ensure_planning_headers(worksheet)

            # Construire la nouvelle ligne (A-L)
            # Colonne J = "ajouté par [username]"
            # Colonne K = Téléphone
            added_by = data.get('added_by', 'Inconnu')
            new_row = [
                "",                                   # A
                data.get('salle', ''),                # B
                data.get('occupant', ''),             # C
                data.get('horaire', ''),              # D
                data.get('date', ''),                 # E
                data.get('accompte', ''),             # F
                data.get('reste_a_payer', ''),        # G
                data.get('prix_location', ''),        # H
                data.get('caution_menage', ''),       # I
                f"ajouté par {added_by}" if added_by else "",  # J
                data.get('telephone', ''),            # K
                "",                                   # L (vide)
            ]

            # TOUJOURS ajouter à la fin du sheet, jamais réutiliser une ligne vide
            all_values, error = self._get_worksheet_values_cached("Planning quotidien")
            if error:
                return False, error

            next_row = len(all_values) + 1
            worksheet.add_rows(1)
            worksheet.update(f'A{next_row}:L{next_row}', [new_row], value_input_option='USER_ENTERED')

            # Invalider le cache pour forcer une relecture fraiche
            self._invalidate_sheet_cache("Planning quotidien")
            return True, f"ligne {next_row}"

        except Exception as e:
            return False, f"Erreur ajout Google Sheets: {str(e)}"

    def check_reservation_conflict(self, salle_name: str, d: date, horaire_str: str) -> list:
        """
        Vérifie si une nouvelle réservation (salle + date + horaire) entre en conflit
        avec les occupations déjà existantes (planning fixe + réservations ponctuelles).
        Retourne une liste d'occupations en conflit (vide si aucun conflit).
        """
        conflits = []
        try:
            debut_new, fin_new = parse_horaire(horaire_str)
        except Exception:
            # Horaire non parseable → on ne peut pas vérifier
            return conflits

        # 1. Conflits avec le planning fixe
        try:
            fixed = self.get_all_fixed_occupations(salle_name, d)
            for occ in fixed:
                debut = occ.get("debut")
                fin = occ.get("fin")
                if not isinstance(debut, time) or not isinstance(fin, time):
                    continue
                if debut == time(0, 0) and fin == time(23, 59) and "warning" in occ:
                    continue
                # Chevauchement d'intervalles
                start_a = self._time_to_minutes(debut_new)
                end_a = self._time_to_minutes(fin_new)
                if end_a <= start_a:
                    end_a += 24 * 60
                start_b = self._time_to_minutes(debut)
                end_b = self._time_to_minutes(fin)
                if end_b <= start_b:
                    end_b += 24 * 60
                if start_a < end_b and start_b < end_a:
                    occ["salle"] = salle_name
                    conflits.append(occ)
        except Exception as e:
            print(f"[check_reservation_conflict] Erreur planning fixe: {e}")

        # 2. Conflits avec les réservations ponctuelles Google Sheets
        try:
            reservations, error = self.get_all_reservations_google(salle_name, d)
            if error:
                print(f"[check_reservation_conflict] Erreur Google Sheets: {error}")
                return conflits
            for occ in reservations:
                debut = occ.get("debut")
                fin = occ.get("fin")
                if not isinstance(debut, time) or not isinstance(fin, time):
                    continue
                if debut == time(0, 0) and fin == time(23, 59) and "warning" in occ:
                    continue
                start_a = self._time_to_minutes(debut_new)
                end_a = self._time_to_minutes(fin_new)
                if end_a <= start_a:
                    end_a += 24 * 60
                start_b = self._time_to_minutes(debut)
                end_b = self._time_to_minutes(fin)
                if end_b <= start_b:
                    end_b += 24 * 60
                if start_a < end_b and start_b < end_a:
                    occ["salle"] = salle_name
                    conflits.append(occ)
        except Exception as e:
            print(f"[check_reservation_conflict] Erreur réservations: {e}")

        return conflits

    def get_all_reservations(self, salle_name: str, d: date) -> tuple:
        """
        Récupère TOUTES les réservations ponctuelles d'une journée (sans filtrer par heure).
        Retourne (results, error) où error est None si succès.
        """
        # Le fallback sur fichier Excel a été supprimé ; les réservations
        # ponctuelles sont désormais lues exclusivement depuis Google Sheets.
        if not self.google_sheet_id:
            return [], "Google Sheets non configuré pour les réservations"

        return self.get_all_reservations_google(salle_name, d)

    def _time_to_minutes(self, t: time) -> int:
        """Convertit un objet time en minutes depuis minuit."""
        return t.hour * 60 + t.minute

    def _find_unprecise_occupations(self, occupations: list) -> list:
        """
        Retourne les occupations dont l'horaire n'est pas numeriquement parseable
        (debut=00:00, fin=23:59 et warning present).
        """
        unprecise = []
        for occ in occupations:
            debut = occ.get("debut")
            fin = occ.get("fin")
            if not isinstance(debut, time) or not isinstance(fin, time):
                continue
            is_full_day_placeholder = (
                debut == time(0, 0) and fin == time(23, 59) and "warning" in occ
            )
            if is_full_day_placeholder:
                unprecise.append({
                    "occupant": occ.get("occupant", "Inconnu"),
                    "horaire": occ.get("horaire", "—"),
                    "source": occ.get("source", "")
                })
        return unprecise

    def _find_overlaps(self, occupations: list) -> list:
        """
        Détecte les occupations qui se chevauchent sur un même créneau horaire.
        Retourne une liste de dicts {first, second, message}.

        Les occupations avec un horaire non parseable (debut=00:00, fin=23:59 et warning)
        ne sont pas comparees car on ne connait pas leur vrai creneau.
        """
        overlaps = []
        if not occupations:
            return overlaps

        segments = []
        for occ in occupations:
            debut = occ.get("debut")
            fin = occ.get("fin")
            if not isinstance(debut, time) or not isinstance(fin, time):
                continue

            # Ignorer les occupations avec horaire non precise (00:00-23:59 + warning)
            is_full_day_placeholder = (
                debut == time(0, 0) and fin == time(23, 59) and "warning" in occ
            )
            if is_full_day_placeholder:
                continue

            start_min = self._time_to_minutes(debut)
            end_min = self._time_to_minutes(fin)
            # Gerer les creneaux overnight
            if end_min <= start_min:
                end_min += 24 * 60

            segments.append({
                "occupant": occ.get("occupant", "Inconnu"),
                "horaire": occ.get("horaire", "—"),
                "source": occ.get("source", ""),
                "start_min": start_min,
                "end_min": end_min,
                "occ": occ
            })

        # Trier par heure de debut
        segments.sort(key=lambda x: x["start_min"])

        for i in range(len(segments)):
            for j in range(i + 1, len(segments)):
                a = segments[i]
                b = segments[j]
                # Chevauchement : a commence avant b finit ET b commence avant a finit
                if a["start_min"] < b["end_min"] and b["start_min"] < a["end_min"]:
                    overlaps.append({
                        "first": a,
                        "second": b,
                        "message": (
                            f"{a['occupant']} ({a['horaire']}) et "
                            f"{b['occupant']} ({b['horaire']}) se chevauchent"
                        )
                    })

        return overlaps

    def get_all_occupations(self, salle_name: str, d: date) -> dict:
        """
        Récupère TOUTES les occupations d'une journée (mode sans heure précise).
        Retourne un dict avec 'error' si la connexion Google Sheets a échoué.
        """
        fixed_occupations = self.get_all_fixed_occupations(salle_name, d)
        reservations, error = self.get_all_reservations(salle_name, d)
        all_occupations = fixed_occupations + reservations

        # Trier par heure de début
        all_occupations.sort(key=lambda x: x["debut"])

        result = {
            "salle": salle_name,
            "date": d,
            "occupations": all_occupations,
            "overlaps": self._find_overlaps(all_occupations),
            "unprecise": self._find_unprecise_occupations(all_occupations)
        }
        if error:
            result["error"] = error
        return result

    def check_availability(self, salle_name: str, d: date, time_requested: time) -> dict:
        """
        Vérifie complètement la disponibilité d'une salle.
        Retourne un dict avec le statut et les détails.
        Ajoute 'error' si la connexion Google Sheets a échoué.
        """
        # Vérifier le planning fixe
        fixed_occupations = self.check_fixed_schedule(salle_name, d, time_requested)

        # Vérifier les réservations ponctuelles
        reservations, error = self.check_reservations(salle_name, d, time_requested)

        # Combiner toutes les occupations
        all_occupations = fixed_occupations + reservations

        # Déterminer le statut
        result = {
            "libre": not all_occupations,
            "salle": salle_name,
            "date": d,
            "heure": time_requested,
            "occupations": all_occupations,
            "overlaps": self._find_overlaps(all_occupations),
            "unprecise": self._find_unprecise_occupations(all_occupations)
        }
        if error:
            result["error"] = error
        return result

    # ═══════════════════════════════════════════════════════════
    # GESTION DES UTILISATEURS (onglet "Utilisateurs" du Sheet)
    # ═══════════════════════════════════════════════════════════
    def get_users_google(self) -> dict:
        """
        Récupère tous les utilisateurs depuis l'onglet 'Utilisateurs' du Google Sheet.
        Retourne un dict compatible streamlit-authenticator:
        {username: {name: ..., password: ...}}
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return {}

        all_values, error = self._get_worksheet_values_cached("Utilisateurs")
        if error or not all_values:
            return {}

        try:
            headers = [str(h).strip().lower() for h in all_values[0]]
            idx_username = headers.index('username') if 'username' in headers else 0
            idx_name = headers.index('name') if 'name' in headers else 1
            idx_password = headers.index('password_hash') if 'password_hash' in headers else 2
            idx_email = headers.index('email') if 'email' in headers else None

            users = {}
            for row in all_values[1:]:
                username = str(row[idx_username]).strip() if len(row) > idx_username else ""
                name = str(row[idx_name]).strip() if len(row) > idx_name else ""
                pwd_hash = str(row[idx_password]).strip() if len(row) > idx_password else ""
                email = ""
                if idx_email is not None and len(row) > idx_email:
                    email = str(row[idx_email]).strip()
                if username and pwd_hash:
                    users[username] = {
                        "name": name or username,
                        "password": pwd_hash,
                        "email": email
                    }
            return users
        except Exception:
            return {}

    def get_notification_emails(self) -> list:
        """
        Récupère la liste des emails valides depuis l'onglet 'Utilisateurs'.
        Utilisé pour l'envoi des notifications.
        """
        users = self.get_users_google()
        emails = []
        for u, data in users.items():
            email = (data.get("email") or "").strip()
            if email and "@" in email:
                emails.append(email)
        return emails

    def get_user_email(self, username: str) -> str:
        """Récupère l'email d'un utilisateur donné."""
        users = self.get_users_google()
        return (users.get(username, {}).get("email") or "").strip()

    def update_user_email_google(self, username: str, email: str) -> tuple:
        """
        Met à jour l'email d'un utilisateur dans l'onglet 'Utilisateurs' (colonne F).
        Retourne (success, error_or_info).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Utilisateurs")
            except gspread.exceptions.WorksheetNotFound:
                return False, "Onglet 'Utilisateurs' introuvable"

            all_values, error = self._get_worksheet_values_cached("Utilisateurs")
            if error:
                return False, error

            for i, row in enumerate(all_values):
                if i == 0:
                    continue  # header
                if len(row) > 0 and str(row[0]).strip() == username:
                    worksheet.update(values=[[email]], range_name=f'F{i+1}', value_input_option='USER_ENTERED')
                    self._invalidate_sheet_cache("Utilisateurs")
                    return True, f"Email de {username} mis à jour"
            return False, f"Utilisateur {username} non trouvé"
        except Exception as e:
            return False, f"Erreur mise à jour email: {str(e)}"

    def add_user_google(self, username: str, name: str, password_hash: str, created_by: str = "", email: str = "") -> tuple:
        """
        Ajoute un nouvel utilisateur dans l'onglet 'Utilisateurs'.
        Retourne (success, error_or_info).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Utilisateurs")
            except gspread.exceptions.WorksheetNotFound:
                # Créer l'onglet s'il n'existe pas
                worksheet = spreadsheet.add_worksheet(title="Utilisateurs", rows=100, cols=6)
                worksheet.update(values=[['username', 'name', 'password_hash', 'created_by', 'created_at', 'email']], range_name='A1:F1')

            # S'assurer que la colonne email (F) existe
            self._ensure_users_headers(worksheet)

            from datetime import datetime
            created_at = datetime.now().strftime("%d/%m/%Y %H:%M")

            all_values, error = self._get_worksheet_values_cached("Utilisateurs")
            if error:
                return False, error

            next_row = len(all_values) + 1
            worksheet.add_rows(1)

            worksheet.update(f'A{next_row}:F{next_row}',
                [[username, name, password_hash, created_by, created_at, email]],
                value_input_option='USER_ENTERED')

            self._invalidate_sheet_cache("Utilisateurs")
            return True, "Compte créé avec succès"
        except Exception as e:
            return False, f"Erreur ajout utilisateur: {str(e)}"

    def update_user_password_google(self, username: str, new_password_hash: str) -> tuple:
        """
        Met à jour le mot de passe d'un utilisateur dans l'onglet 'Utilisateurs'.
        Retourne (success, error_or_info).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Utilisateurs")
            except gspread.exceptions.WorksheetNotFound:
                return False, "Onglet 'Utilisateurs' introuvable"

            all_values, error = self._get_worksheet_values_cached("Utilisateurs")
            if error:
                return False, error

            for i, row in enumerate(all_values):
                if i == 0:
                    continue  # header
                if len(row) > 0 and str(row[0]).strip() == username:
                    worksheet.update(f'C{i+1}', [[new_password_hash]], value_input_option='USER_ENTERED')
                    self._invalidate_sheet_cache("Utilisateurs")
                    return True, f"Mot de passe de {username} mis à jour"

            return False, f"Utilisateur {username} non trouvé"
        except Exception as e:
            return False, f"Erreur mise à jour mot de passe: {str(e)}"

    def delete_user_google(self, username: str) -> tuple:
        """
        Supprime un utilisateur (ligne entière) dans l'onglet 'Utilisateurs'.
        Retourne (success, error_or_info).
        """
        if not self.google_sheet_id or not GSPREAD_AVAILABLE:
            return False, "Google Sheets non configuré"

        try:
            client, error = self._get_google_client()
            if error:
                return False, error
            if not client:
                return False, "Client Google Sheets non initialisé"

            spreadsheet = client.open_by_key(self.google_sheet_id)
            try:
                worksheet = spreadsheet.worksheet("Utilisateurs")
            except gspread.exceptions.WorksheetNotFound:
                return False, "Onglet 'Utilisateurs' introuvable"

            all_values, error = self._get_worksheet_values_cached("Utilisateurs")
            if error:
                return False, error

            for i, row in enumerate(all_values):
                if i == 0:
                    continue  # header
                if len(row) > 0 and str(row[0]).strip() == username:
                    worksheet.delete_rows(i + 1)
                    self._invalidate_sheet_cache("Utilisateurs")
                    return True, f"Utilisateur {username} supprimé"

            return False, f"Utilisateur {username} non trouvé"
        except Exception as e:
            return False, f"Erreur suppression utilisateur: {str(e)}"
